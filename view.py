# -------------------------------------------------------------------------- #
# IMPORTAÇÕES

# tkinter (messagebox)
from tkinter.messagebox import showinfo

# sqlite3
import sqlite3

# openpyxl
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Font

# time (para anexar em nome dos arquivos)
from time import strftime


# -------------------------------------------------------------------------- #
# CONEXÃO


connection = sqlite3.connect('data.db')


# -------------------------------------------------------------------------- #
# INSERÇÕES


def insert_data(data):
    """Cuida da inserção de registros."""
    with connection:
        cursor = connection.cursor()
        insertion = '''
            insert into inventario(
                nome, descricao, unidades,
                marca, data, valor, serie,
                local_da_imagem
            ) values(?, ?, ? , ?, ?, ?, ?, ?)
        '''
        cursor.execute(insertion, data)


# -------------------------------------------------------------------------- #
# REMOÇÕES


def delete_data(id):
    """Cuida de deletar registros."""
    with connection:
        cursor = connection.cursor()
        deletion = 'delete from inventario where id = ?'
        cursor.execute(deletion, id)


def drop_table():
    """Cuida de apagar a tabela toda."""
    with connection:
        cursor = connection.cursor()
        cursor.execute('delete from inventario')


# -------------------------------------------------------------------------- #
# EDIÇÕES/MANIPULAÇÕES


def update_data(data):
    """Cuida de atualizar registros."""
    with connection:
        cursor = connection.cursor()
        update = '''
            update inventario set
            nome = ?, descrição = ?,
            unidades = ?, marca = ?,
            data = ?, valor = ?,
            serie = ?, local_da_imagem = ?
            where id = ?
        '''
        cursor.execute(update, data)


# -------------------------------------------------------------------------- #
# CONSULTAS


def show_data():
    """Cuida de mostrar os registros da tabela."""
    data = list()
    with connection:
        cursor = connection.cursor()
        query = 'select * from inventario'
        cursor.execute(query)

        rows = cursor.fetchall()

        for row in rows:
            data.append(row)

    return data


def show_record(id):
    """Cuida de mostrar um único registro."""
    with connection:
        cursor = connection.cursor()
        query = 'select * from inventario where id = ?'
        cursor.execute(query, id)

        row = cursor.fetchone()

    return row


# -------------------------------------------------------------------------- #
# EXPORTA/IMPORTA TABELA PARA/DE EXCEL


def export_to_excel():
    """Cuida de salvar a tabela em Excel."""
    with connection:
        cursor = connection.cursor()

        # Cabeçalho da tabela, menos o id
        table_headers = [
            'Nome', 'Descrição', 'Unidades',
            'Marca/Modelo', 'Data de Aquisição',
            'Valor', 'Serial',
            'Local da Imagem'
        ]

        # Pega os registros da tabela
        datalist = show_data()

        # Cria uma nova lista menos o id
        new_list = list()
        for record in datalist:
            new_list.append(record[1:9])

        # Insere no inicio da lista o cabeçalho
        new_list.insert(0, table_headers)

        # Cria a planilha e a 'folha'
        wb = Workbook()
        ws = wb.active

        # Coloca os dados e o
        # cabeçalho na planilha
        for row in new_list:
            ws.append(row)

        # Coloca o cabeçalho, com
        # a fonte em negrito.
        ft = Font(bold=True)
        for row in ws['A1:H1']:
            for cell in row:
                cell.font = ft


        '''
        Aqui, tento colocar as células na planilha com uma largura aceitável
        de se vêr. Não fica perfeito mas, pelo menos as células não
        ficam todas espremidas kk.
        '''
        for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            max_width = 0

            for row_number in range(1, ws.max_row + 1):
                if len(str(ws[f'{letter}{row_number}'].value)) > max_width:
                    max_width = len(str(ws[f'{letter}{row_number}'].value))

            ws.column_dimensions[letter].width = max_width + 1


        # Salva a planilha.
        wb.save(f"inventario-{strftime('%s')}.xlsx")

        showinfo(
            'Sucesso',
            'Tabela exportada para arquivo Excel com sucesso.'
        )


def import_from_excel(excel_file):
    """Cuida de importar os dados da Planilha."""
    with connection:
        cursor = connection.cursor()

        # Abre o excel para lêr
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        '''
        Cria a lista para o sqlite. min_row=2, porque não queremos
        o cabeçalho da planilha e list() em tudo o que é lugar
        porque não suporto tuplas kk
        '''
        datalist = list()
        for row in ws.iter_rows(min_row=2, values_only=True):
            datalist.append(list(row))

        '''
        Aqui, crio new_datalist() e temp_list(), para trocar os items
        que forem do tipo None por uma string vazia ''. Faço isto para
        não dar erro em main.py, por exemplo ao carregar no botão ver item.
        Deve ser MUITO ineficiente fazer isto desta forma, mas ainda não me
        sinto á vontade com os outros métodos que encontrei na internet,
        como expressões lambda e map etc..
        '''
        new_datalist = list()
        for inner_list in datalist:
            temp_list = list()
            for item in inner_list:
                if item is None:
                    item = str('')
                temp_list.append(item)
            new_datalist.append(temp_list)

        # Coloca na tabela
        cursor.executemany(
            '''
            insert into inventario(
                nome, descricao, unidades,
                marca, data, valor, serie,
                local_da_imagem
            ) values(?, ?, ?, ?, ?, ?, ?, ?)
            ''', new_datalist
        )

        showinfo(
            'Sucesso',
            'Dados importados de arquivo Excel para tabela com sucesso'
        )

        return
